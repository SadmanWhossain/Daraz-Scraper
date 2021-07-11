from selenium import webdriver
from openpyxl import workbook, load_workbook
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
import time
import os

from PyQt5 import QtCore, QtGui, QtWidgets

class Ui_DarazScraper(object):
    def setupUi(self, DarazScraper):
        DarazScraper.setObjectName("DarazScraper")
        DarazScraper.resize(500, 200)
        font = QtGui.QFont()
        font.setFamily("Segoe UI Semibold")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        DarazScraper.setFont(font)
        self.label = QtWidgets.QLabel(DarazScraper)
        self.label.setGeometry(QtCore.QRect(30, 70, 121, 16))
        self.label.setObjectName("label")
        self.lineEdit = QtWidgets.QLineEdit(DarazScraper)
        self.lineEdit.setGeometry(QtCore.QRect(170, 60, 291, 31))
        self.lineEdit.setObjectName("lineEdit")
        self.pushButton = QtWidgets.QPushButton(DarazScraper)
        self.pushButton.setGeometry(QtCore.QRect(154, 122, 191, 41))
        self.pushButton.setObjectName("pushButton")

        self.retranslateUi(DarazScraper)
        QtCore.QMetaObject.connectSlotsByName(DarazScraper)

    def retranslateUi(self, DarazScraper):
        _translate = QtCore.QCoreApplication.translate
        DarazScraper.setWindowTitle(_translate("DarazScraper", "Daraz Scraper"))
        self.label.setText(_translate("DarazScraper", "Search Keyword"))
        self.pushButton.setText(_translate("DarazScraper", "Collect"))

class scraping:
    def execution(self, keyword):
        wb = load_workbook(keyword + '.xlsx')

        ws = wb[keyword]
        # ws = wb.create_sheet("vans")

        # print(wb.sheetnames)
        path = "G:/Project/Daraz-Scraper/chromedriver.exe"
        driver = webdriver.Chrome(path)

        fake_name_websites = driver.get("https://www.daraz.com.bd/")

        searchbox = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//input[@type='search']")))
        searchbox.clear()
        k = "sunglass"
        searchbox.send_keys(k)
        searchbutton = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(),'SEARCH')]"))).click()

        links = []

        products = driver.find_elements_by_xpath("//div[@data-tracking='product-card']//a")
        for product in products:
            links.append(product.get_property('href'))
        print(len(links))
        print(links)




if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    DarazScraper = QtWidgets.QWidget()
    ui = Ui_DarazScraper()
    ui.setupUi(DarazScraper)
    DarazScraper.show()
    sys.exit(app.exec_())

